'''
Author: Jadedever
Date: 2022-05-03 15:21:23
LastEditors: Jadedever
LastEditTime: 2022-05-03 15:24:08
FilePath: /Python_Homework/lesson_7/read.py
Description:  

Copyright (c) 2022 by Jadedever, All Rights Reserved. 
'''
from openpyxl import load_workbook

wb=load_workbook('考勤统计.xlsx')
sheet=wb['8 月考勤统计表']

print(wb.sheetnames)

print(sheet['A1'].value)

# 打印所有单元格的值
for row in sheet.rows:
  for cell in row:
    print(cell.value)


